"""
qr_id_generator.py
==================
Generates QR-code-enhanced ID cards for Kannada Koota Luxembourg (KKL).

Reads member data directly from the KKL membership Excel file (.xlsx).

─────────────────────────────────────────────────────────────────────────────
CARD LAYOUT
─────────────────────────────────────────────────────────────────────────────
  ┌─────────────────────────────────────────┐
  │ Member:      <right-justified name>     │
  │ ID Number:   <right-justified id>       │
  │ Category:    <right-justified category> │
  │ Validity:    <right-justified date>     │
  │                                         │
  │         [ QR code with KKL logo ]       │
  └─────────────────────────────────────────┘

  - Yellow-to-red vertical gradient background
  - All four legend labels left-justified at a fixed x position
  - All four field values right-justified at the right edge of the card
  - Name truncated to 30 characters maximum, breaking at a word boundary
  - Large QR code centred below the text block
  - KKL logo embedded in the centre of the QR code (ERROR_CORRECT_H)

─────────────────────────────────────────────────────────────────────────────
EXCEL COLUMNS USED
─────────────────────────────────────────────────────────────────────────────
  'Category'                     – Family or Individual
  'Full name of the main member' – Primary member name
  'family member count'          – Number of members (informational)
  'Member ID'                    – Contains 'KKL2026-' prefix
  'Valid Until'                  – Expiry date stored as Excel datetime

─────────────────────────────────────────────────────────────────────────────
RUNTIME PROMPTS
─────────────────────────────────────────────────────────────────────────────
  1. Starting sequential number  → appended to KKL2026- prefix
                                    e.g. 1 → KKL2026-001, KKL2026-002 …
  2. Validity date override      → optional; uses Excel date if skipped

─────────────────────────────────────────────────────────────────────────────
OUTPUT
─────────────────────────────────────────────────────────────────────────────
  <QR_CODE_DIR>/<ID_Number>_qr.png  – standalone QR code image with logo
  <ID_CARD_DIR>/<ID_Number>_id.png  – finished, print-ready ID card

─────────────────────────────────────────────────────────────────────────────
DEPENDENCIES
─────────────────────────────────────────────────────────────────────────────
  pip install qrcode[pil] Pillow openpyxl
"""

import logging
import os
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
import qrcode
from PIL import Image, ImageDraw, ImageFont

# =============================================================================
# LOGGING SETUP
# Timestamps + log level prefix on every line written to stdout.
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger(__name__)


# =============================================================================
# CONFIGURATION
# All file paths default to the folder containing this script so the project
# works immediately after a `git clone` without editing any source.
# Override any path by setting the corresponding environment variable before
# running the script, e.g.:
#     export QR_EXCEL_PATH=/data/members_2026.xlsx
# =============================================================================

# Absolute path to the folder that contains this script file.
# Used as the base for all relative default paths below.
_SCRIPT_DIR: Path = Path(__file__).resolve().parent

# ── Input ────────────────────────────────────────────────────────────────────

# Path to the KKL membership Excel workbook (.xlsx).
# The script reads the ACTIVE sheet; column names must match exactly.
EXCEL_PATH: Path = Path(os.getenv("QR_EXCEL_PATH", _SCRIPT_DIR / "members.xlsx"))

# ── Output ───────────────────────────────────────────────────────────────────

# Directory where intermediate QR code PNG files are saved.
QR_CODE_DIR: Path = Path(os.getenv("QR_CODE_DIR", _SCRIPT_DIR / "qr_codes"))

# Directory where the finished ID card PNG files are saved.
ID_CARD_DIR: Path = Path(os.getenv("ID_CARD_DIR", _SCRIPT_DIR / "id_cards"))

# ── Branding ─────────────────────────────────────────────────────────────────

# Path to the KKL circular logo PNG.
# The logo is resized and composited into the centre of every QR code.
# If the file is missing, QR codes are generated without a logo (with a warning).
LOGO_PATH: Path = Path(os.getenv("QR_LOGO_PATH", _SCRIPT_DIR / "logo.png"))

# Fixed prefix for every Membership ID.  The sequential number is appended
# at runtime, e.g. KKL2026-001, KKL2026-002, …
MEMBERSHIP_PREFIX: str = "KKL2026-"

# ── Card dimensions ───────────────────────────────────────────────────────────

# Card canvas size in pixels – portrait orientation to match the KKL sample.
CARD_WIDTH:  int = 550
CARD_HEIGHT: int = 750

# ── Date ─────────────────────────────────────────────────────────────────────

# Fallback validity date used when the Excel cell is empty or unreadable.
VALID_UNTIL_DEFAULT: str = "2026-12-31"

# ── Typography ────────────────────────────────────────────────────────────────

# Path to a TrueType Bold font.
# Default points to DejaVu Sans Bold which ships with GitHub Codespaces.
# On Windows, change to e.g. "C:/Windows/Fonts/arialbd.ttf" (Arial Bold).
FONT_PATH: str = os.getenv(
    "QR_FONT_PATH",
    "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
)

# Maximum number of characters allowed in the Member name field on the card.
# Names longer than this are truncated at the nearest word boundary so no
# word is split in the middle.
NAME_MAX_CHARS: int = 30

# ── QR code ───────────────────────────────────────────────────────────────────

# Side length (pixels) of the QR code image pasted onto the card.
QR_RENDER_SIZE: int = 460

# The KKL logo covers this fraction of the QR code's width and height.
# Must stay at or below 0.30 so scanners can still decode the code despite
# the logo obscuring the centre modules.
LOGO_FRACTION: float = 0.28

# ── Colour palette ────────────────────────────────────────────────────────────

# RGB tuple for the top of the background gradient (bright yellow).
GRADIENT_TOP:    tuple = (255, 213,   0)

# RGB tuple for the bottom of the background gradient (deep red).
GRADIENT_BOTTOM: tuple = (180,  20,   0)

# Colour used for all text drawn on the card.
COLOUR_TEXT: str = "black"

# ── Layout constants ──────────────────────────────────────────────────────────

# Horizontal padding (pixels) from the left edge of the card to the labels.
TEXT_PADDING_LEFT: int = 18

# Horizontal padding (pixels) from the RIGHT edge of the card to the values.
# Values are right-aligned so their rightmost pixel is at CARD_WIDTH - this.
TEXT_PADDING_RIGHT: int = 18

# Vertical gap (pixels) between the top of the card and the first text line.
TEXT_PADDING_TOP: int = 20

# Vertical gap (pixels) between consecutive text lines.
LINE_SPACING: int = 14

# Vertical gap (pixels) between the last text line and the top of the QR code.
QR_TOP_MARGIN: int = 20

# =============================================================================
# EXCEL COLUMN NAMES
# These strings must match the header row in members.xlsx exactly
# (case-sensitive, including spaces).
# =============================================================================
COL_CATEGORY:    str = "Category"
COL_NAME:        str = "Full name of the main member"
COL_COUNT:       str = "family member count"
COL_MEMBER_ID:   str = "Member ID"
COL_VALID_UNTIL: str = "Valid Until"


# =============================================================================
# HELPER: NAME TRUNCATION
# =============================================================================

def truncate_name(name: str, max_chars: int = NAME_MAX_CHARS) -> str:
    """
    Shorten *name* to at most *max_chars* characters, breaking only at a
    word boundary so no word is split in the middle.

    Algorithm:
        1. If the name fits within *max_chars*, return it unchanged.
        2. Otherwise, take the substring up to *max_chars* characters and
           find the last space within that substring.
        3. Strip everything from that space onwards and add "…" (ellipsis)
           to indicate truncation.
        4. If the substring contains no space at all (one very long word),
           hard-truncate at *max_chars* and add "…".

    Examples (max_chars=30):
        "Ranjana Ramachandra Bhat"      → "Ranjana Ramachandra Bhat"  (24 chars, no change)
        "Subramanyam Umamaheshwaran X"  → "Subramanyam…"              (truncated at word)

    Args:
        name:      The raw member name string.
        max_chars: Maximum number of characters permitted (default NAME_MAX_CHARS).

    Returns:
        The original name if it fits, otherwise a word-boundary-truncated
        string ending with "…".
    """
    if len(name) <= max_chars:
        return name                         # name fits – return unchanged

    # Work within the first max_chars characters
    truncated = name[:max_chars]

    last_space = truncated.rfind(" ")       # find the rightmost space

    if last_space > 0:
        # Break at the last complete word
        return truncated[:last_space] + "…"
    else:
        # No space found – hard cut at max_chars
        return truncated + "…"


# =============================================================================
# HELPER: GRADIENT BACKGROUND
# =============================================================================

def _make_gradient(width: int, height: int,
                   top: tuple, bottom: tuple) -> Image.Image:
    """
    Create a vertical linear gradient image.

    Each horizontal scan line is painted with a colour that is linearly
    interpolated between *top* (at y=0) and *bottom* (at y=height-1).

    Args:
        width:  Canvas width in pixels.
        height: Canvas height in pixels.
        top:    RGB tuple for the colour at the very top of the image.
        bottom: RGB tuple for the colour at the very bottom of the image.

    Returns:
        A new RGB PIL Image whose pixels form the requested gradient.
    """
    base = Image.new("RGB", (width, height))
    draw = ImageDraw.Draw(base)

    for y in range(height):
        # ratio goes from 0.0 (top) to 1.0 (bottom)
        ratio = y / (height - 1)

        # Linearly interpolate each RGB channel independently
        r = int(top[0] + (bottom[0] - top[0]) * ratio)
        g = int(top[1] + (bottom[1] - top[1]) * ratio)
        b = int(top[2] + (bottom[2] - top[2]) * ratio)

        # Paint a single-pixel-tall horizontal line in the interpolated colour
        draw.line([(0, y), (width, y)], fill=(r, g, b))

    return base


# =============================================================================
# HELPER: FONT LOADER
# =============================================================================

def _load_fonts(path: str) -> tuple:
    """
    Load the TrueType font at *path* in three sizes for use on the ID card.

    Font sizes:
        font_large  – 28 pt  used for Member name and ID Number
        font_medium – 24 pt  used for Category and Validity
        font_small  – 20 pt  reserved for future use / footnotes

    Graceful fallback:
        If the TTF file cannot be found or opened, PIL's built-in bitmap font
        is used instead.  The bitmap font is the same object for all three
        size variables (it does not scale), so text sizes will look identical
        on the card.  A WARNING is logged so the operator knows to fix the
        font path.

    Args:
        path: Filesystem path to a .ttf (TrueType) font file.

    Returns:
        Tuple of (font_large, font_medium, font_small) PIL font objects.
    """
    try:
        font_large  = ImageFont.truetype(path, 28)
        font_medium = ImageFont.truetype(path, 24)
        font_small  = ImageFont.truetype(path, 20)
        logger.debug("Loaded TrueType font: %s", path)

    except (OSError, IOError):
        logger.warning(
            "Font file not found at '%s'. "
            "Falling back to PIL built-in bitmap font – "
            "text sizing on cards will be approximate.",
            path,
        )
        default     = ImageFont.load_default()
        font_large  = default
        font_medium = default
        font_small  = default

    return font_large, font_medium, font_small


# =============================================================================
# HELPER: TEXT BLOCK RENDERER
# =============================================================================

def _draw_text_block(
    draw: ImageDraw.ImageDraw,
    lines: list[tuple[str, str, object]],
    start_y: int,
    card_width: int,
) -> int:
    """
    Render the four-line member detail block onto *draw*.

    Layout rules applied here:
      • Labels  → left-justified at TEXT_PADDING_LEFT pixels from the left edge.
      • Values  → right-justified so their rightmost pixel is at
                  (card_width - TEXT_PADDING_RIGHT).
      • Each line occupies its natural text height plus LINE_SPACING pixels.

    Args:
        draw:       PIL ImageDraw object bound to the card canvas.
        lines:      List of (label, value, font) tuples, one per text row.
        start_y:    Vertical pixel position for the top of the first line.
        card_width: Total card width in pixels (used to compute right edge).

    Returns:
        The y-coordinate immediately below the last drawn line (i.e. the y
        position available for the next element, such as the QR code).
    """
    y = start_y

    for label, value, font in lines:

        # ── Draw label (LEFT-justified) ───────────────────────────────────
        draw.text(
            (TEXT_PADDING_LEFT, y),
            label,
            fill=COLOUR_TEXT,
            font=font,
        )

        # ── Draw value (RIGHT-justified) ──────────────────────────────────
        # textbbox returns (left, top, right, bottom) in pixels for the given
        # string rendered at position (0, 0).  We use the width (right - left)
        # to calculate where the text must START so it ends at the right margin.
        value_bbox  = draw.textbbox((0, 0), value, font=font)
        value_width = value_bbox[2] - value_bbox[0]  # pixel width of value string
        value_x     = card_width - TEXT_PADDING_RIGHT - value_width

        draw.text(
            (value_x, y),
            value,
            fill=COLOUR_TEXT,
            font=font,
        )

        # ── Advance y by line height + spacing ────────────────────────────
        # Use the label's bounding box for line height (label is always present;
        # value could theoretically be empty).
        label_bbox   = draw.textbbox((0, 0), label, font=font)
        line_height  = label_bbox[3] - label_bbox[1]
        y            += line_height + LINE_SPACING

    return y   # caller uses this as the top of the QR code


# =============================================================================
# QR CODE GENERATION WITH LOGO OVERLAY
# =============================================================================

def generate_qr_code(data: str, id_number: str, output_dir: Path) -> Path:
    """
    Encode *data* as a QR code, overlay the KKL logo in the centre, and
    save the result as a PNG file.

    Why ERROR_CORRECT_H?
        The QR specification allows up to ~30 % of the code surface to be
        obscured when using the highest error correction level (H).  The KKL
        logo covers LOGO_FRACTION (28 %) of the QR area, so H is required to
        guarantee reliable scanning despite the logo obstruction.

    Logo placement:
        1. The QR image is converted to RGBA so alpha compositing works.
        2. A white filled ellipse (slightly larger than the logo) is painted
           in the centre to give the logo a clean circular white halo.
        3. The logo is pasted over the halo using its own alpha channel as
           the transparency mask.

    Args:
        data:       String payload to encode in the QR code.
                    Typically contains member name, ID, category, and validity.
        id_number:  Membership ID used to construct the output filename.
        output_dir: Directory path where the QR PNG is written.

    Returns:
        Path object pointing to the saved QR code PNG file.

    Raises:
        OSError: If the output file cannot be written.
    """
    # ── Build the QR code matrix ─────────────────────────────────────────────
    qr = qrcode.QRCode(
        version=1,              # start at the smallest version; auto-grows
        box_size=10,            # pixels per individual QR module (dot)
        border=4,               # quiet-zone width in modules (spec minimum = 4)
        error_correction=qrcode.constants.ERROR_CORRECT_H,  # ~30 % recovery
    )
    qr.add_data(data)
    qr.make(fit=True)           # automatically select the smallest version that fits

    # Convert to a PIL image and scale to the desired render size
    qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGBA")
    qr_img = qr_img.resize((QR_RENDER_SIZE, QR_RENDER_SIZE), Image.LANCZOS)

    # ── Overlay the KKL logo ─────────────────────────────────────────────────
    if LOGO_PATH.exists():

        # Scale logo to LOGO_FRACTION of the QR code size
        logo_size = int(QR_RENDER_SIZE * LOGO_FRACTION)

        with Image.open(LOGO_PATH).convert("RGBA") as logo_raw:
            logo = logo_raw.resize((logo_size, logo_size), Image.LANCZOS)

        # Centre coordinates for the logo within the QR image
        logo_x = (QR_RENDER_SIZE - logo_size) // 2
        logo_y = (QR_RENDER_SIZE - logo_size) // 2

        # Create a transparent RGBA layer the same size as the QR image,
        # paint a white filled ellipse behind the logo area for a clean halo
        halo_layer = Image.new("RGBA", (QR_RENDER_SIZE, QR_RENDER_SIZE), (0, 0, 0, 0))
        halo_draw  = ImageDraw.Draw(halo_layer)
        halo_pad   = 6   # pixels of white halo extending beyond the logo edge

        halo_draw.ellipse(
            [
                logo_x - halo_pad,
                logo_y - halo_pad,
                logo_x + logo_size + halo_pad,
                logo_y + logo_size + halo_pad,
            ],
            fill=(255, 255, 255, 255),   # solid white
        )

        # Merge the white halo onto the QR code using alpha compositing
        qr_img = Image.alpha_composite(qr_img, halo_layer)

        # Paste the logo using its own alpha channel as the transparency mask
        qr_img.paste(logo, (logo_x, logo_y), mask=logo)
        logger.debug("KKL logo overlaid on QR code for ID %s.", id_number)

    else:
        logger.warning(
            "Logo file not found at '%s'. "
            "QR code for ID %s generated without logo.",
            LOGO_PATH, id_number,
        )

    # ── Save to disk ─────────────────────────────────────────────────────────
    qr_final = qr_img.convert("RGB")          # PNG does not require alpha
    qr_path  = output_dir / f"{id_number}_qr.png"
    qr_final.save(qr_path)
    logger.debug("QR code saved: %s", qr_path)
    return qr_path


# =============================================================================
# ID CARD CREATION
# =============================================================================

def create_id_card(
    name: str,
    id_number: str,
    category: str,
    valid_until: str,
    qr_dir: Path,
    card_dir: Path,
) -> Path:
    """
    Composite a single KKL ID card and save it as a PNG.

    Steps performed:
        1. Paint a yellow-to-red vertical gradient background.
        2. Truncate the member name to NAME_MAX_CHARS characters at a word
           boundary to prevent overflow into the QR code area.
        3. Format the validity date as DD/MM/YYYY for display.
        4. Draw four text rows with LEFT-justified labels and RIGHT-justified
           values using _draw_text_block().
        5. Generate the QR code (with embedded logo) via generate_qr_code().
        6. Paste the QR code centred horizontally, below the text block.
        7. Save the finished card to *card_dir*.

    Text rows on the card:
        Member:     <name, max 30 chars>
        ID Number:  <KKL2026-NNN>
        Category:   <Family | Individual>
        Validity:   <DD/MM/YYYY>

    Args:
        name:        Raw member name from Excel (may be longer than 30 chars).
        id_number:   Constructed Membership ID (e.g. KKL2026-001).
        category:    'Family' or 'Individual' from the Excel Category column.
        valid_until: Expiry date string in YYYY-MM-DD format.
        qr_dir:      Directory for the intermediate QR PNG.
        card_dir:    Directory for the finished ID card PNG.

    Returns:
        Path to the saved ID card PNG.

    Raises:
        OSError: If either the QR file or the card file cannot be written.
    """
    # ── 1. Gradient background ────────────────────────────────────────────────
    card = _make_gradient(CARD_WIDTH, CARD_HEIGHT, GRADIENT_TOP, GRADIENT_BOTTOM)
    draw = ImageDraw.Draw(card)

    # ── 2. Load fonts ─────────────────────────────────────────────────────────
    font_large, font_medium, font_small = _load_fonts(FONT_PATH)

    # ── 3. Truncate name to NAME_MAX_CHARS at a word boundary ─────────────────
    display_name = truncate_name(name, NAME_MAX_CHARS)

    # ── 4. Format validity date as DD/MM/YYYY ─────────────────────────────────
    try:
        d = date.fromisoformat(valid_until)
        validity_display = d.strftime("%d/%m/%Y")
    except ValueError:
        # If date parsing fails, use the raw string (better than crashing)
        validity_display = valid_until
        logger.warning("Could not parse date '%s' for ID %s.", valid_until, id_number)

    # ── 5. Define the four text rows ──────────────────────────────────────────
    # Each tuple: (label string, value string, font object)
    # Labels are the fixed printed legends; values are the member-specific data.
    # Note: "Membership ID" has been renamed to "ID Number" per requirements.
    lines = [
        ("Member:",    display_name,      font_large),   # row 1 – member name
        ("ID Number:", id_number,         font_large),   # row 2 – KKL2026-NNN
        ("Category:",  category,          font_medium),  # row 3 – Family/Individual
        ("Validity:",  validity_display,  font_medium),  # row 4 – DD/MM/YYYY
    ]

    # ── 6. Draw text block (labels left, values right) ────────────────────────
    text_bottom_y = _draw_text_block(
        draw=draw,
        lines=lines,
        start_y=TEXT_PADDING_TOP,
        card_width=CARD_WIDTH,
    )

    # ── 7. Build QR code payload ──────────────────────────────────────────────
    # Keep payload concise; most QR scanners display limited text.
    qr_payload = (
        f"Member:{name}\n"          # use raw (untruncated) name in QR data
        f"ID:{id_number}\n"
        f"Category:{category}\n"
        f"Validity:{validity_display}"
    )
    qr_path = generate_qr_code(qr_payload, id_number, qr_dir)

    # ── 8. Paste QR code centred horizontally below the text block ────────────
    with Image.open(qr_path) as qr_img:
        # Re-scale to QR_RENDER_SIZE in case generate_qr_code produced a
        # different size (should not happen, but defensive is better).
        qr_resized = qr_img.resize((QR_RENDER_SIZE, QR_RENDER_SIZE), Image.LANCZOS)

    qr_x = (CARD_WIDTH - QR_RENDER_SIZE) // 2   # horizontally centred
    qr_y = text_bottom_y + QR_TOP_MARGIN          # directly below text block

    card.paste(qr_resized, (qr_x, qr_y))

    # ── 9. Save finished card ─────────────────────────────────────────────────
    output_path = card_dir / f"{id_number}_id.png"
    card.save(output_path)
    logger.info("Created ID card: %s", output_path)
    return output_path


# =============================================================================
# EXCEL PROCESSING
# =============================================================================

def process_excel(
    excel_path: Path,
    qr_dir: Path,
    card_dir: Path,
    start_sequence: int,
    valid_until_override: str | None,
) -> tuple[int, int]:
    """
    Iterate over the KKL membership Excel workbook and generate one ID card
    per valid data row.

    Processing rules:
        • Rows with an empty 'Full name of the main member' cell are skipped
          with a WARNING log entry.
        • Rows with an empty 'Category' cell are skipped with a WARNING.
        • Valid Until is taken from the Excel cell unless the operator
          provided a command-line override at startup.
        • The Membership ID sequential counter increments only when a card
          is successfully created; skipped rows do NOT consume a number.

    Args:
        excel_path:           Path to the .xlsx membership workbook.
        qr_dir:               Directory for QR code PNG output files.
        card_dir:             Directory for ID card PNG output files.
        start_sequence:       First integer appended to MEMBERSHIP_PREFIX
                              (e.g. 1 → KKL2026-001).
        valid_until_override: When not None, this date string (YYYY-MM-DD)
                              is used for ALL cards, overriding the Excel
                              cell values.

    Returns:
        Tuple of (success_count, skip_count) where:
            success_count – number of ID cards successfully written to disk.
            skip_count    – number of rows skipped due to missing data or errors.
    """
    # ── Open workbook ─────────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active   # always process the first (active) sheet

    # ── Map column names to zero-based indices ────────────────────────────────
    # The first row is treated as the header row.
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    logger.info("Excel header columns: %s", headers)

    # Abort early if any required column is absent from the header row
    required_cols = [COL_CATEGORY, COL_NAME, COL_VALID_UNTIL]
    for col in required_cols:
        if col not in headers:
            logger.error(
                "Required column '%s' not found in the Excel header row. "
                "Check that the column name matches exactly (case-sensitive). "
                "Aborting.",
                col,
            )
            sys.exit(1)

    # Build a dict: column_name → column_index (0-based)
    col_idx = {col_name: idx for idx, col_name in enumerate(headers)}

    # ── Iterate over data rows ────────────────────────────────────────────────
    success = 0
    skipped = 0
    seq     = start_sequence   # current sequential number for ID generation

    for row_number, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True),   # skip header (row 1)
        start=2,                                       # display row numbers from 2
    ):
        # ── Extract cell values ───────────────────────────────────────────────
        name     = row[col_idx[COL_NAME]]
        category = row[col_idx[COL_CATEGORY]]
        val_raw  = row[col_idx[COL_VALID_UNTIL]]

        # ── Validate: skip rows with missing name ─────────────────────────────
        if not name or not str(name).strip():
            logger.warning("Row %d: 'Full name of the main member' is empty – skipping.", row_number)
            skipped += 1
            continue

        # ── Validate: skip rows with missing category ─────────────────────────
        if not category or not str(category).strip():
            logger.warning("Row %d: 'Category' is empty – skipping.", row_number)
            skipped += 1
            continue

        # Normalise to plain strings with surrounding whitespace removed
        name     = str(name).strip()
        category = str(category).strip()

        # ── Resolve the Valid Until date ──────────────────────────────────────
        if valid_until_override:
            # Operator supplied an override at startup: use for every card
            valid_until = valid_until_override

        elif isinstance(val_raw, datetime):
            # openpyxl returns Excel date cells as Python datetime objects
            valid_until = val_raw.strftime("%Y-%m-%d")

        elif isinstance(val_raw, str) and val_raw.strip():
            # Occasionally dates are stored as plain text in the Excel cell
            valid_until = val_raw.strip()

        else:
            # Cell is empty or an unrecognised type: fall back to the default
            valid_until = VALID_UNTIL_DEFAULT
            logger.warning(
                "Row %d: 'Valid Until' cell is empty or unreadable – "
                "using default date %s.",
                row_number, VALID_UNTIL_DEFAULT,
            )

        # ── Build the Membership ID for this row ──────────────────────────────
        # Zero-pad the sequence number to 3 digits: 1 → 001, 12 → 012, etc.
        id_number = f"{MEMBERSHIP_PREFIX}{seq:03d}"

        # ── Generate the ID card ──────────────────────────────────────────────
        try:
            create_id_card(
                name=name,
                id_number=id_number,
                category=category,
                valid_until=valid_until,
                qr_dir=qr_dir,
                card_dir=card_dir,
            )
            success += 1
            seq     += 1   # advance the counter ONLY on a successful card

        except Exception:
            # Log the full traceback but continue processing remaining rows
            logger.exception(
                "Row %d: unexpected error while creating card for '%s' (ID=%s).",
                row_number, name, id_number,
            )
            skipped += 1

    return success, skipped


# =============================================================================
# INTERACTIVE PROMPTS
# =============================================================================

def _prompt_start_sequence() -> int:
    """
    Ask the operator for the starting sequential number for Membership IDs.

    The number is zero-padded to 3 digits and appended to MEMBERSHIP_PREFIX:
        Input 1  → KKL2026-001, KKL2026-002, KKL2026-003 …
        Input 13 → KKL2026-013, KKL2026-014, KKL2026-015 …

    Keeps looping until a valid positive integer is entered.
    Pressing Enter without typing a value defaults to 1.

    Returns:
        Positive integer to use as the first sequential ID number.
    """
    while True:
        raw = input(
            f"\nEnter starting sequence number for {MEMBERSHIP_PREFIX} [1]: "
        ).strip()

        # Empty input → use default of 1
        if not raw:
            print("Using default starting number: 1")
            return 1

        # Parse and validate the input
        try:
            n = int(raw)
            if n < 1:
                raise ValueError("Sequence number must be at least 1.")
            return n
        except ValueError:
            print("  ✗  Please enter a positive whole number (e.g. 1, 13, 100).")


def _prompt_valid_until_override() -> str | None:
    """
    Optionally ask the operator to override the Valid Until date for ALL cards.

    The Excel workbook already contains a Valid Until date for each member row.
    This prompt lets the operator replace that date across the entire batch
    (useful if the Excel has not been updated yet for the new membership year).

    Pressing Enter without typing a date keeps the per-row Excel dates.

    Returns:
        A validated YYYY-MM-DD date string if the operator entered one, or
        None to use the dates stored in the Excel workbook.
    """
    print(
        "\nValid Until dates are normally read from the Excel file."
        "\nPress Enter to use the Excel dates for each member, "
        "or type a date to override ALL cards."
    )
    raw = input("Override date [YYYY-MM-DD] or Enter to skip: ").strip()

    # Empty input → keep Excel dates
    if not raw:
        print("Using Valid Until dates from Excel.")
        return None

    # Validate the supplied date string; keep prompting until it is valid
    while True:
        try:
            date.fromisoformat(raw)   # raises ValueError if format is wrong
            return raw
        except ValueError:
            raw = input(
                "  ✗  Invalid format. Enter YYYY-MM-DD "
                "or press Enter to use Excel dates: "
            ).strip()
            if not raw:
                print("Using Valid Until dates from Excel.")
                return None


# =============================================================================
# ENTRY POINT
# =============================================================================

def main() -> None:
    """
    Main entry point for the KKL ID Card Generator.

    Execution order:
        1. Display startup banner.
        2. Prompt for starting sequence number.
        3. Prompt for optional Valid Until date override.
        4. Warn if the logo file is missing (non-fatal).
        5. Abort if the Excel file is missing (fatal).
        6. Create output directories if they do not already exist.
        7. Process every data row in the Excel workbook.
        8. Report the final success/skip counts.

    Exit codes:
        0 – all rows processed successfully.
        1 – fatal error (Excel file not found; aborts before processing).
        2 – partial success (at least one row was skipped or failed).
    """
    print("\n" + "=" * 48)
    print("   KKL ID Card Generator")
    print("=" * 48)

    # ── Step 2: Prompt for starting sequence number ───────────────────────────
    start_sequence = _prompt_start_sequence()

    # ── Step 3: Prompt for optional date override ─────────────────────────────
    valid_until_override = _prompt_valid_until_override()

    # Log the confirmed settings before starting batch processing
    logger.info("Starting Membership ID : %s%03d", MEMBERSHIP_PREFIX, start_sequence)
    logger.info(
        "Valid Until            : %s",
        valid_until_override if valid_until_override else "from Excel per row",
    )

    # ── Step 4: Logo check (non-fatal warning) ────────────────────────────────
    if not LOGO_PATH.exists():
        logger.warning(
            "Logo file not found at '%s'. "
            "Cards will be generated without the KKL logo in the QR code. "
            "To add the logo, place logo.png in the same folder as this script.",
            LOGO_PATH,
        )

    # ── Step 5: Excel file check (fatal) ─────────────────────────────────────
    if not EXCEL_PATH.exists():
        logger.error(
            "Excel membership file not found: %s  "
            "Place members.xlsx in the same folder as this script, "
            "or set the QR_EXCEL_PATH environment variable.",
            EXCEL_PATH,
        )
        sys.exit(1)

    # ── Step 6: Create output directories ────────────────────────────────────
    QR_CODE_DIR.mkdir(parents=True, exist_ok=True)
    ID_CARD_DIR.mkdir(parents=True, exist_ok=True)
    logger.info("QR codes output  →  %s", QR_CODE_DIR)
    logger.info("ID cards output  →  %s", ID_CARD_DIR)

    # ── Step 7: Process all rows ──────────────────────────────────────────────
    logger.info("Reading membership data from: %s", EXCEL_PATH)
    success, skipped = process_excel(
        excel_path=EXCEL_PATH,
        qr_dir=QR_CODE_DIR,
        card_dir=ID_CARD_DIR,
        start_sequence=start_sequence,
        valid_until_override=valid_until_override,
    )

    # ── Step 8: Final summary ─────────────────────────────────────────────────
    print("\n" + "=" * 48)
    logger.info(
        "Finished.  %d card(s) created successfully,  %d row(s) skipped.",
        success, skipped,
    )
    print("=" * 48)

    # Exit with code 2 if any rows were skipped (useful for CI pipelines)
    if skipped:
        sys.exit(2)


# =============================================================================
# SCRIPT ENTRY GUARD
# =============================================================================
if __name__ == "__main__":
    main()
